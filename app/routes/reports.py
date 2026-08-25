import io

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from flask import Blueprint, render_template, send_file

from app.database import get_db

bp = Blueprint("reports", __name__)


@bp.get("/reports")
def reports():
    with get_db() as db:
        with db.cursor() as cur:
            cur.execute(
                """
                SELECT COALESCE(translated_status, 'Unknown') AS status, COUNT(*) AS total
                FROM inbound_responses
                GROUP BY translated_status
                ORDER BY total DESC
                """
            )
            by_code = cur.fetchall()

            cur.execute(
                """
                SELECT from_number,
                       COALESCE(translated_status, 'Unknown') AS status,
                       COUNT(*) AS total,
                       MAX(received_at) AS last_received
                FROM inbound_responses
                GROUP BY from_number, translated_status
                ORDER BY from_number, status
                """
            )
            by_address = cur.fetchall()

            cur.execute(
                """
                SELECT address_ref, body, ozeki_msg_id, status, delivery_status, sent_at, campaign_id
                FROM outbound_messages
                ORDER BY sent_at DESC
                LIMIT 50
                """
            )
            outbound = cur.fetchall()

            cur.execute(
                """
                SELECT c.id, c.body, c.created_at, c.recipient_count,
                       COUNT(DISTINCT o.id) AS sent_count,
                       SUM(CASE WHEN o.delivery_status = 'DELIVERED' THEN 1 ELSE 0 END) AS delivered_count,
                       COUNT(DISTINCT ir.from_number) AS responded_count
                FROM campaigns c
                LEFT JOIN outbound_messages o  ON o.campaign_id = c.id
                LEFT JOIN inbound_responses ir ON ir.campaign_id = c.id
                GROUP BY c.id
                ORDER BY c.created_at DESC
                LIMIT 20
                """
            )
            campaigns = cur.fetchall()

    return render_template(
        "report.html",
        by_code=by_code,
        by_address=by_address,
        outbound=outbound,
        campaigns=campaigns,
    )


@bp.get("/reports/export")
def export_excel():
    with get_db() as db:
        with db.cursor() as cur:
            cur.execute(
                """
                SELECT from_number, raw_message, response_code,
                       COALESCE(translated_status, 'Unknown') AS translated_status,
                       received_at
                FROM inbound_responses
                ORDER BY received_at DESC
                """
            )
            responses = cur.fetchall()

            cur.execute(
                """
                SELECT address_ref, body, ozeki_msg_id, status, delivery_status, sent_at, campaign_id
                FROM outbound_messages
                ORDER BY sent_at DESC
                LIMIT 500
                """
            )
            outbound = cur.fetchall()

    wb = openpyxl.Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1A3A5C")
    center = Alignment(horizontal="center")

    # Sheet 1: Inbound responses
    ws1 = wb.active
    ws1.title = "Inbound Responses"
    ws1.append(["From Number", "Message", "Response Code", "Status", "Received At"])
    for cell in ws1[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
    for row in responses:
        ws1.append([
            row["from_number"],
            row["raw_message"],
            row["response_code"],
            row["translated_status"],
            str(row["received_at"]) if row["received_at"] else "",
        ])
    ws1.column_dimensions["A"].width = 18
    ws1.column_dimensions["B"].width = 30
    ws1.column_dimensions["C"].width = 16
    ws1.column_dimensions["D"].width = 20
    ws1.column_dimensions["E"].width = 22

    # Sheet 2: Outbound messages
    ws2 = wb.create_sheet("Outbound Messages")
    ws2.append(["Recipient", "Message", "Message ID", "Send Status", "Delivery Status", "Sent At", "Campaign ID"])
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
    for row in outbound:
        ws2.append([
            row["address_ref"],
            row["body"],
            row["ozeki_msg_id"] or "",
            row["status"] or "",
            row["delivery_status"] or "",
            str(row["sent_at"]) if row["sent_at"] else "",
            row["campaign_id"] or "",
        ])
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 40
    ws2.column_dimensions["C"].width = 16
    ws2.column_dimensions["D"].width = 14
    ws2.column_dimensions["E"].width = 16
    ws2.column_dimensions["F"].width = 22
    ws2.column_dimensions["G"].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="openvox_responses.xlsx",
    )
